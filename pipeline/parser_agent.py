"""
pipeline/parser_agent.py
------------------------
Parser Agent: reads paper_info.xlsx to locate the replication code (Stata DO or R),
parses the regression specification, and returns a typed PaperSpec dict consumed by
every downstream agent.

Usage
-----
    python pipeline/parser_agent.py Paper_Meyer2024
    python pipeline/parser_agent.py --all
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Literal, Optional, TypedDict

import openpyxl

# ---------------------------------------------------------------------------
# PaperSpec TypedDict
# ---------------------------------------------------------------------------

class PaperSpec(TypedDict):
    # Identity
    paper_id: str
    paper_dir: str
    title: str
    source_do_file: Optional[str]
    source_data_file: Optional[str]
    source_r_file: Optional[str]
    replication_code_type: Literal["stata", "r", "python", "none"]
    # Regression spec
    estimator: str        # OLS / FE / RE / IV / Logit / Probit / GLS / GMM / HLM / Other
    dependent_var: str
    key_independent_vars: list
    control_vars: list
    fixed_effects: list
    cluster_var: Optional[str]
    sample_restrictions: list
    interaction_terms: list
    instrumental_vars: list
    # Published result
    published_coef: Optional[float]
    published_se: Optional[float]
    published_significance: Optional[str]
    # Audit flags
    parse_confidence: Literal["high", "medium", "low"]
    flags: list
    manual_review_required: bool
    raw_regression_command: Optional[str]


# ---------------------------------------------------------------------------
# xlsx column index constants (1-based)
# ---------------------------------------------------------------------------
_COL = {
    "paper_short_name":     1,
    "source_folder":        2,
    "do_file":              3,
    "data_file":            4,
    "estimator":            5,
    "depvar":               6,
    "main_coef":            7,
    "indepvars":            8,
    "absorb":               9,
    "cluster":              10,
    "published_coef_main":  11,
    "key_vars":             12,
    "aux_var":              13,
    "status":               14,
    "notes":                15,
}


# ---------------------------------------------------------------------------
# Path resolution helpers
# ---------------------------------------------------------------------------

def _resolve_source_folder(source_folder: str, ra_task_root: Path) -> Optional[Path]:
    """Resolve a (possibly truncated) source_folder value to an actual directory.

    The xlsx stores source_folder strings that may end with '...' due to cell
    width limits, or may differ from the actual folder name.  Strategy:
    1. Exact match.
    2. Strip trailing '.' characters → prefix.  Then accept any directory d where
       prefix.startswith(d) OR d.startswith(prefix).
       Tie-break: longest matching directory name.
    """
    exact = ra_task_root / source_folder
    if exact.is_dir():
        return exact

    # Build prefix by stripping trailing dots
    prefix = source_folder.rstrip(".")
    candidates = []
    try:
        for d in ra_task_root.iterdir():
            if not d.is_dir():
                continue
            dname = d.name
            if prefix.startswith(dname) or dname.startswith(prefix):
                candidates.append(d)
    except FileNotFoundError:
        return None

    if not candidates:
        return None
    # Prefer longest match
    return max(candidates, key=lambda d: len(d.name))


def _read_paper_xlsx(paper_dir: Path) -> dict:
    """Read paper_info.xlsx and return a flat dict of column values."""
    xlsx_path = paper_dir / "paper_info.xlsx"
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    row = [c.value for c in ws[2]]
    return dict(zip(headers, row))


# ---------------------------------------------------------------------------
# Text pre-processing for Stata DO files
# ---------------------------------------------------------------------------

def _preprocess_do(text: str) -> str:
    """Strip comments, join continuation lines."""
    # 1. Remove /* ... */ block comments (including multi-line)
    text = re.sub(r'/\*.*?\*/', '', text, flags=re.DOTALL)

    # 2. Join /// continuation lines
    while True:
        new = re.sub(r'///[^\n]*\n[ \t]*', ' ', text)
        if new == text:
            break
        text = new

    # 3. Process line by line: strip // comments and * comment lines
    cleaned = []
    for line in text.split('\n'):
        stripped = line.strip()
        # Skip * comment lines
        if stripped.startswith('*'):
            continue
        # Strip // line comments (but not ///  — already handled)
        line = re.sub(r'(?<!/)//(?!/).*', '', line)
        cleaned.append(line)

    return '\n'.join(cleaned)


# ---------------------------------------------------------------------------
# Macro resolution
# ---------------------------------------------------------------------------

def _collect_macros(lines: list[str]) -> tuple[dict, dict]:
    """Collect global and local macro definitions."""
    globals_: dict[str, str] = {}
    locals_: dict[str, str] = {}

    global_pat = re.compile(
        r'^\s*global\s+(\w+)\s*(?:=\s*"([^"]*)"|\s*=\s*(\S+)|"([^"]*)"|(.*?))$',
        re.IGNORECASE,
    )
    local_pat = re.compile(
        r'^\s*local\s+(\w+)\s*(?:=\s*"([^"]*)"|\s*=\s*(\S+)|"([^"]*)"|(.*?))$',
        re.IGNORECASE,
    )

    for line in lines:
        gm = global_pat.match(line)
        if gm:
            name = gm.group(1)
            val = next((v for v in gm.groups()[1:] if v is not None), "")
            globals_[name] = val.strip()
            continue
        lm = local_pat.match(line)
        if lm:
            name = lm.group(1)
            val = next((v for v in lm.groups()[1:] if v is not None), "")
            locals_[name] = val.strip()

    return globals_, locals_


def _expand_macros(text: str, globals_: dict, locals_: dict) -> tuple[str, list[str]]:
    """Expand $global and `local' references; return expanded text and any unresolved flags."""
    flags: list[str] = []

    # Expand globals: $macroname or ${macroname}
    def replace_global(m: re.Match) -> str:
        name = m.group(1) or m.group(2)
        if name in globals_:
            return globals_[name]
        flags.append(f"unresolved macro: ${name}")
        return m.group(0)

    text = re.sub(r'\$\{(\w+)\}|\$(\w+)', replace_global, text)

    # Expand locals: `macroname'
    def replace_local(m: re.Match) -> str:
        name = m.group(1)
        if name in locals_:
            return locals_[name]
        flags.append(f"unresolved macro: `{name}'")
        return m.group(0)

    text = re.sub(r'`(\w+)\'', replace_local, text)

    return text, flags


# ---------------------------------------------------------------------------
# Stata variable name stripping
# ---------------------------------------------------------------------------

_STATA_PREFIX = re.compile(
    r'^(?:(?:[LFDlfd]\d*\.)|(?:[cCiI]\.)|(?:[LF]\d+\.))*'
)

def _strip_stata_prefix(varname: str) -> str:
    """Remove Stata time-series/factor operator prefixes (i., l., f., L2., etc.)."""
    return _STATA_PREFIX.sub('', varname).strip()


# ---------------------------------------------------------------------------
# Interaction term expansion
# ---------------------------------------------------------------------------

def _expand_interactions(token: str) -> tuple[list[str], list[str]]:
    """Parse a token like 'i.state_id#c.fyear' or 'Post##Treat'.

    Returns (variable_names, interaction_label_list).
    """
    # ## operator
    if '##' in token:
        parts = token.split('##')
        vars_ = [_strip_stata_prefix(p) for p in parts if p]
        return vars_, [token]
    # # operator
    if '#' in token:
        parts = token.split('#')
        vars_ = [_strip_stata_prefix(p) for p in parts if p]
        return vars_, [token]
    return [], []


# ---------------------------------------------------------------------------
# Stata regression command finder & parser
# ---------------------------------------------------------------------------

_REGCMDS = re.compile(
    r'^\s*(?:qui(?:etly)?\s+)?(?:xi\s*:?\s*)?'
    r'(reghdfe|areg|regress\b|reg\b|xtreg\b|ivregress\b|ivreg2\b|xtivreg\b|'
    r'xtlogit\b|xtprobit\b|logit\b|probit\b|xtgls\b|xtmixed\b|mixed\b|glm\b)',
    re.IGNORECASE,
)

_ESTIMATOR_MAP = {
    'regress': 'OLS', 'reg': 'OLS',
    'areg': 'FE', 'reghdfe': 'FE',
    'xtreg': None,          # depends on ,fe / ,re option
    'ivregress': 'IV', 'ivreg2': 'IV', 'xtivreg': 'IV',
    'logit': 'Logit', 'xtlogit': 'Logit',
    'probit': 'Probit', 'xtprobit': 'Probit',
    'xtgls': 'GLS',
    'mixed': 'HLM', 'xtmixed': 'HLM',
    'glm': 'Other',
}


def _parse_reg_command(cmd: str) -> dict:
    """Parse a single (cleaned) Stata regression command string into components."""
    result: dict = {
        'raw': cmd,
        'estimator': 'OLS',
        'dv': '',
        'indepvars': [],
        'fixed_effects': [],
        'cluster_var': None,
        'sample_restrictions': [],
        'interaction_terms': [],
        'instrumental_vars': [],
    }

    # Remove xi: prefix
    cmd_work = re.sub(r'^(?:qui(?:etly)?\s+)?(?:xi\s*:?\s*)?', '', cmd.strip(), flags=re.IGNORECASE)

    # Extract command keyword
    kw_match = re.match(r'^(\w+)', cmd_work)
    if not kw_match:
        return result
    kw = kw_match.group(1).lower()
    after_kw = cmd_work[len(kw):].strip()

    # Map estimator
    base_est = _ESTIMATOR_MAP.get(kw, 'OLS')
    result['estimator'] = base_est or 'OLS'   # will refine for xtreg below

    # For ivregress, skip method keyword (2sls / liml / gmm)
    if kw in ('ivregress', 'ivreg2'):
        after_kw = re.sub(r'^(?:2sls|liml|gmm)\s+', '', after_kw, flags=re.IGNORECASE)

    # Split on comma to separate varlist from options
    comma_idx = _find_option_comma(after_kw)
    if comma_idx >= 0:
        varlist_part = after_kw[:comma_idx].strip()
        options_part = after_kw[comma_idx + 1:].strip()
    else:
        varlist_part = after_kw.strip()
        options_part = ''

    # Extract "if" condition from varlist
    if_match = re.search(r'\bif\b', varlist_part, re.IGNORECASE)
    if if_match:
        cond = varlist_part[if_match.end():].strip()
        # strip trailing "in ..." if present
        in_m = re.search(r'\bin\b', cond, re.IGNORECASE)
        if in_m:
            cond = cond[:in_m.start()].strip()
        result['sample_restrictions'].append(cond)
        varlist_part = varlist_part[:if_match.start()].strip()

    # Handle IV instruments: (endog = instrs)
    iv_match = re.search(r'\(([^)]+)=([^)]*)\)', varlist_part)
    if iv_match:
        # endog var(s) and instruments
        endog_vars = iv_match.group(1).split()
        instr_vars = iv_match.group(2).split()
        result['instrumental_vars'] = [_strip_stata_prefix(v) for v in instr_vars if v]
        # Remove IV block from varlist
        varlist_part = (varlist_part[:iv_match.start()] + ' ' +
                        ' '.join(endog_vars) + ' ' +
                        varlist_part[iv_match.end():]).strip()

    tokens = varlist_part.split()
    if not tokens:
        return result

    # First token is DV
    dv = _strip_stata_prefix(tokens[0])
    result['dv'] = dv

    # Remaining tokens are indepvars (skip i.varname factor indicators — keep them for FE)
    indep_raw = tokens[1:]
    indepvars: list[str] = []
    fe_from_i: list[str] = []
    interaction_terms: list[str] = []

    for tok in indep_raw:
        # Handle interaction tokens
        inter_vars, inter_labels = _expand_interactions(tok)
        if inter_labels:
            interaction_terms.extend(inter_labels)
            # Add component vars to indepvars (stripped)
            for v in inter_vars:
                if v and v not in indepvars:
                    indepvars.append(v)
            continue

        # i.varname → fixed effect via indepvar notation
        if re.match(r'^i\.', tok, re.IGNORECASE):
            varname = _strip_stata_prefix(tok)
            if varname not in fe_from_i:
                fe_from_i.append(varname)
            continue

        cleaned = _strip_stata_prefix(tok)
        if cleaned and cleaned not in indepvars:
            indepvars.append(cleaned)

    result['indepvars'] = indepvars
    result['fixed_effects'] = fe_from_i
    result['interaction_terms'] = interaction_terms

    # --- Parse options ---
    opts = options_part

    # xtreg ,fe / ,re
    if kw == 'xtreg':
        if re.search(r'\bfe\b', opts, re.IGNORECASE):
            result['estimator'] = 'FE'
        elif re.search(r'\bre\b', opts, re.IGNORECASE):
            result['estimator'] = 'RE'
        else:
            result['estimator'] = 'FE'  # default for xtreg

    # absorb() or a() — fixed effects
    absorb_match = re.search(r'\b(?:absorb|a)\(([^)]+)\)', opts, re.IGNORECASE)
    if absorb_match:
        absorb_content = absorb_match.group(1)
        # Split on whitespace and +
        raw_fes = re.split(r'[\s+]+', absorb_content)
        for fe_token in raw_fes:
            fe_token = fe_token.strip()
            if not fe_token:
                continue
            # Handle interaction FEs like i.sic_2#i.fyear
            if '#' in fe_token:
                parts = fe_token.split('#')
                for p in parts:
                    v = _strip_stata_prefix(p.strip())
                    if v and v not in result['fixed_effects']:
                        result['fixed_effects'].append(v)
            else:
                v = _strip_stata_prefix(fe_token)
                if v and v not in result['fixed_effects']:
                    result['fixed_effects'].append(v)

    # cluster: cl(var), cluster(var), vce(cluster var)
    cl_match = (
        re.search(r'\bcl\((\w+)\)', opts, re.IGNORECASE) or
        re.search(r'\bcluster\((\w+)\)', opts, re.IGNORECASE) or
        re.search(r'\bvce\(\s*cluster\s+(\w+)\s*\)', opts, re.IGNORECASE)
    )
    if cl_match:
        result['cluster_var'] = cl_match.group(1)

    return result


def _find_option_comma(varlist_str: str) -> int:
    """Find the index of the first unparenthesized comma (option separator)."""
    depth = 0
    for i, ch in enumerate(varlist_str):
        if ch == '(':
            depth += 1
        elif ch == ')':
            depth -= 1
        elif ch == ',' and depth == 0:
            return i
    return -1


# ---------------------------------------------------------------------------
# "Most complete" regression selector
# ---------------------------------------------------------------------------

def _count_unique_vars(parsed: dict) -> int:
    """Count unique non-trivial independent variables."""
    return len(set(parsed.get('indepvars', [])))


def _find_main_regression(do_text: str) -> Optional[dict]:
    """Find and parse the 'most complete' regression command in cleaned DO text."""
    lines = do_text.split('\n')
    candidates: list[dict] = []
    loop_depth = 0

    for line in lines:
        stripped = line.strip()
        # Track loop depth
        if re.match(r'^\s*(?:forval(?:ues)?|foreach)\b', stripped, re.IGNORECASE):
            loop_depth += 1
        if re.match(r'^\s*\}', stripped):
            loop_depth = max(0, loop_depth - 1)

        if loop_depth > 0:
            continue  # skip commands inside loops

        if _REGCMDS.match(stripped):
            parsed = _parse_reg_command(stripped)
            candidates.append(parsed)

    if not candidates:
        return None

    # Select most complete (most indepvars), tie-break = last
    best = candidates[0]
    best_count = _count_unique_vars(best)
    for c in candidates[1:]:
        cnt = _count_unique_vars(c)
        if cnt >= best_count:
            best = c
            best_count = cnt

    return best


# ---------------------------------------------------------------------------
# DO file parser: public entry
# ---------------------------------------------------------------------------

def _parse_do_file(do_path: Path) -> PaperSpec:
    """Parse a Stata DO file and return a (partial) PaperSpec."""
    try:
        text = do_path.read_text(encoding='utf-8')
    except UnicodeDecodeError:
        text = do_path.read_text(encoding='latin-1')

    cleaned = _preprocess_do(text)
    lines = cleaned.split('\n')
    globals_, locals_ = _collect_macros(lines)
    expanded, macro_flags = _expand_macros(cleaned, globals_, locals_)

    reg = _find_main_regression(expanded)

    flags: list[str] = list(macro_flags)
    spec_dict: dict = {
        'replication_code_type': 'stata',
        'source_do_file': str(do_path),
        'source_r_file': None,
        'raw_regression_command': reg['raw'] if reg else None,
        'estimator': reg['estimator'] if reg else '',
        'dependent_var': reg['dv'] if reg else '',
        'key_independent_vars': [],
        'control_vars': [],
        'fixed_effects': reg['fixed_effects'] if reg else [],
        'cluster_var': reg['cluster_var'] if reg else None,
        'sample_restrictions': reg['sample_restrictions'] if reg else [],
        'interaction_terms': reg['interaction_terms'] if reg else [],
        'instrumental_vars': reg['instrumental_vars'] if reg else [],
    }

    if reg:
        ivs = reg['indepvars']
        if ivs:
            spec_dict['key_independent_vars'] = [ivs[0]]
            spec_dict['control_vars'] = ivs[1:]
            flags.append("key_independent_vars heuristic — verify manually")
        else:
            flags.append("no independent variables found — verify manually")
    else:
        flags.append("no regression command found — verify manually")

    # Confidence
    dv_ok = bool(spec_dict['dependent_var'])
    iv_ok = len(spec_dict['key_independent_vars']) + len(spec_dict['control_vars']) >= 2
    macro_ok = not any('unresolved macro' in f for f in flags)

    if dv_ok and iv_ok and macro_ok:
        confidence: Literal["high", "medium", "low"] = "high"
    elif not dv_ok:
        confidence = "low"
    else:
        confidence = "medium"

    spec_dict['parse_confidence'] = confidence
    spec_dict['flags'] = flags
    spec_dict['manual_review_required'] = confidence != 'high' or bool(flags)

    return spec_dict  # type: ignore[return-value]


# ---------------------------------------------------------------------------
# R file parser
# ---------------------------------------------------------------------------

_R_REGCMDS = re.compile(
    r'\b(lm|felm|lmer|glm|plm|systemfit)\s*\(',
    re.IGNORECASE,
)


def _parse_r_formula(formula: str) -> tuple[str, list[str]]:
    """Extract DV and IVs from a simple R formula string 'DV ~ IV1 + IV2 + ...'."""
    parts = formula.split('~', 1)
    if len(parts) != 2:
        return '', []
    dv = parts[0].strip()
    rhs = parts[1].strip()
    # For felm: formula | FE | instruments | cluster — take only the first section
    rhs = rhs.split('|')[0].strip()
    # Split by + and -
    ivs = [t.strip().lstrip('-').strip() for t in re.split(r'[+\-]', rhs) if t.strip()]
    # Remove R-style operators
    ivs = [re.sub(r'^[Ii]\(', '', v).rstrip(')') for v in ivs if v and v != '1' and v != '0']
    return dv, ivs


def _parse_r_file(r_path: Path) -> dict:
    """Parse an R script and return a partial spec dict."""
    text = r_path.read_text(encoding='utf-8', errors='replace')
    flags = ['R script — verify translation to Python']

    # Check for systemfit
    if re.search(r'\bsystemfit\s*\(', text, re.IGNORECASE):
        flags.append('SUR model in R — may need systemfit/GLS equivalent in Python')

    best_dv = ''
    best_ivs: list[str] = []
    best_fe: list[str] = []
    best_cluster: Optional[str] = None
    best_raw: Optional[str] = None
    best_estimator = 'OLS'

    for m in _R_REGCMDS.finditer(text):
        func = m.group(1).lower()
        # Grab everything from ( to balanced )
        start = m.end() - 1
        depth = 0
        end = start
        for i in range(start, min(start + 2000, len(text))):
            if text[i] == '(':
                depth += 1
            elif text[i] == ')':
                depth -= 1
                if depth == 0:
                    end = i + 1
                    break
        call_content = text[start:end]

        # Extract first formula argument
        formula_match = re.search(r'([A-Za-z_\.][A-Za-z0-9_\.]*\s*~[^,\)]+)', call_content)
        if not formula_match:
            continue

        raw_formula = formula_match.group(1).strip()
        dv, ivs = _parse_r_formula(raw_formula)
        if not dv:
            continue

        estimator = 'OLS'
        fe: list[str] = []
        cluster: Optional[str] = None

        if func == 'felm':
            estimator = 'FE'
            # felm(formula | FE | instruments | cluster)
            sections = call_content.lstrip('(').split('|')
            if len(sections) >= 2:
                fe_section = sections[1].strip().rstrip(')')
                fe = [v.strip() for v in re.split(r'[+\s]+', fe_section) if v.strip() and v.strip() != '0']
            if len(sections) >= 4:
                cluster_section = sections[3].strip()
                # Strip closing paren from the call and any trailing args like ", data=df)"
                cluster_clean = re.split(r'[,\s]', cluster_section.strip('() \n'))[0].strip()
                if cluster_clean and cluster_clean != '0':
                    cluster = cluster_clean
        elif func in ('lmer', 'xtmixed'):
            estimator = 'HLM'
        elif func == 'plm':
            estimator = 'FE'
        elif func == 'systemfit':
            estimator = 'GLS'

        if len(ivs) >= len(best_ivs):
            best_dv, best_ivs, best_fe, best_cluster, best_raw, best_estimator = (
                dv, ivs, fe, cluster, raw_formula, estimator
            )

    key_ivs = [best_ivs[0]] if best_ivs else []
    controls = best_ivs[1:] if len(best_ivs) > 1 else []

    return {
        'replication_code_type': 'r',
        'source_r_file': str(r_path),
        'source_do_file': None,
        'raw_regression_command': best_raw,
        'estimator': best_estimator,
        'dependent_var': best_dv,
        'key_independent_vars': key_ivs,
        'control_vars': controls,
        'fixed_effects': best_fe,
        'cluster_var': best_cluster,
        'sample_restrictions': [],
        'interaction_terms': [],
        'instrumental_vars': [],
        'parse_confidence': 'medium',
        'flags': flags,
        'manual_review_required': True,
    }


# ---------------------------------------------------------------------------
# No-code spec
# ---------------------------------------------------------------------------

def _no_code_spec(paper_id: str, paper_dir: str) -> dict:
    return {
        'replication_code_type': 'none',
        'source_do_file': None,
        'source_r_file': None,
        'source_data_file': None,
        'raw_regression_command': None,
        'estimator': '',
        'dependent_var': '',
        'key_independent_vars': [],
        'control_vars': [],
        'fixed_effects': [],
        'cluster_var': None,
        'sample_restrictions': [],
        'interaction_terms': [],
        'instrumental_vars': [],
        'parse_confidence': 'low',
        'flags': ['no replication code — spec must be entered manually in paper_info.xlsx'],
        'manual_review_required': True,
    }


# ---------------------------------------------------------------------------
# PDF coefficient extraction (optional, best-effort)
# ---------------------------------------------------------------------------

def _extract_pdf_coef(pdf_path: Path) -> tuple[Optional[float], Optional[str]]:
    """Try to extract the main published coefficient from a PDF.  Returns (coef, significance)."""
    try:
        import pdfplumber  # type: ignore
        with pdfplumber.open(pdf_path) as pdf:
            text = '\n'.join(p.extract_text() or '' for p in pdf.pages[:20])
    except ImportError:
        try:
            from pypdf import PdfReader  # type: ignore
            reader = PdfReader(str(pdf_path))
            text = '\n'.join(p.extract_text() or '' for p in reader.pages[:20])
        except ImportError:
            return None, None
    except Exception:
        return None, None

    # Look for coefficient patterns near significance stars
    pattern = re.compile(
        r'(-?\d+\.\d+)\s*(\*{1,3})'
    )
    matches = pattern.findall(text)
    if not matches:
        return None, None
    # Return first match with highest significance
    best_coef, best_sig = matches[0]
    for coef_str, sig in matches:
        if len(sig) > len(best_sig):
            best_coef, best_sig = coef_str, sig
    try:
        return float(best_coef), sig
    except ValueError:
        return None, None


# ---------------------------------------------------------------------------
# Save spec to disk
# ---------------------------------------------------------------------------

def _save_spec(spec: dict, paper_dir: Path) -> None:
    """Write spec.json, update paper_info.xlsx, append to log."""
    # 1. Write spec.json
    spec_path = paper_dir / "spec.json"
    with open(spec_path, 'w', encoding='utf-8') as f:
        json.dump(spec, f, indent=2)

    # 2. Update paper_info.xlsx
    xlsx_path = paper_dir / "paper_info.xlsx"
    if xlsx_path.exists():
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        # Update fields that the parser knows about
        ws.cell(row=2, column=_COL['estimator']).value = spec.get('estimator') or None
        ws.cell(row=2, column=_COL['depvar']).value = spec.get('dependent_var') or None
        ws.cell(row=2, column=_COL['indepvars']).value = (
            ' '.join(spec.get('key_independent_vars', []) + spec.get('control_vars', []))
            or None
        )
        absorb_val = spec.get('fixed_effects', [])
        ws.cell(row=2, column=_COL['absorb']).value = ' + '.join(absorb_val) if absorb_val else None
        ws.cell(row=2, column=_COL['cluster']).value = spec.get('cluster_var')
        key_ivs = spec.get('key_independent_vars', [])
        main_coef = key_ivs[0] if key_ivs else None
        ws.cell(row=2, column=_COL['main_coef']).value = main_coef
        wb.save(xlsx_path)

    # 3. Append to log
    logs_dir = paper_dir.parent.parent / 'logs'
    logs_dir.mkdir(exist_ok=True)
    paper_id = spec.get('paper_id', paper_dir.name)
    log_path = logs_dir / f"{paper_id}_log.md"
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    control_count = len(spec.get('control_vars', []))
    flags_str = '; '.join(spec.get('flags', [])) or '—'
    log_entry = (
        f"\n## Parser Agent — {timestamp}\n"
        f"- Confidence: {spec.get('parse_confidence', '?')}\n"
        f"- Estimator: {spec.get('estimator', '?')}\n"
        f"- DV: {spec.get('dependent_var', '?')}\n"
        f"- Key IVs: {spec.get('key_independent_vars', [])}\n"
        f"- Controls count: {control_count}\n"
        f"- FE: {spec.get('fixed_effects', [])}\n"
        f"- Cluster: {spec.get('cluster_var', 'None')}\n"
        f"- Flags: {flags_str}\n"
        f"- Manual review: {spec.get('manual_review_required', True)}\n"
    )
    with open(log_path, 'a', encoding='utf-8') as f:
        f.write(log_entry)


# ---------------------------------------------------------------------------
# Main entry: parse_paper
# ---------------------------------------------------------------------------

def parse_paper(
    paper_id: str,
    papers_root: Path,
    ra_task_root: Optional[Path] = None,
) -> PaperSpec:
    """Parse a paper's replication code and return a PaperSpec.

    Parameters
    ----------
    paper_id:
        Directory name under papers_root (e.g. 'Paper_Meyer2024').
    papers_root:
        Path to the 'papers/' directory.
    ra_task_root:
        Path to 'RA Missing data task/'.  Defaults to papers_root parent / 'RA Missing data task'.
    """
    paper_dir = papers_root / paper_id
    if ra_task_root is None:
        ra_task_root = papers_root.parent / 'RA Missing data task'

    xlsx_data = _read_paper_xlsx(paper_dir)
    source_folder = xlsx_data.get('source_folder') or ''
    do_file = (xlsx_data.get('do_file') or '').strip()
    data_file = xlsx_data.get('data_file')

    # Resolve source folder
    src_dir: Optional[Path] = _resolve_source_folder(source_folder, ra_task_root)

    # Determine what kind of code we have
    is_none = do_file.upper().startswith('NONE')
    is_r_hint = re.search(r'R script[:\s]+(\S+\.R)', do_file, re.IGNORECASE)

    partial: dict

    if not is_none and src_dir is not None:
        # Standard Stata DO file
        do_path = src_dir / do_file
        if do_path.exists():
            partial = _parse_do_file(do_path)
        else:
            partial = _no_code_spec(paper_id, str(paper_dir))
            partial['flags'].append(f"DO file not found at {do_path}")
    elif is_r_hint and src_dir is not None:
        # Explicit R file hint in do_file field
        r_filename = is_r_hint.group(1)
        r_path = src_dir / r_filename
        if r_path.exists():
            partial = _parse_r_file(r_path)
        else:
            partial = _no_code_spec(paper_id, str(paper_dir))
            partial['flags'].append(f"R file not found at {r_path}")
    elif is_none and src_dir is not None:
        # Check for R files in source folder
        r_files = list(src_dir.glob('*.R')) + list(src_dir.glob('*.r'))
        if r_files:
            partial = _parse_r_file(r_files[0])
        else:
            partial = _no_code_spec(paper_id, str(paper_dir))
    else:
        partial = _no_code_spec(paper_id, str(paper_dir))

    # PDF coefficient extraction (optional)
    pdf_paths = list(paper_dir.glob('*.pdf'))
    if src_dir and not pdf_paths:
        pdf_paths = list(src_dir.glob('*.pdf'))

    published_coef: Optional[float] = None
    published_significance: Optional[str] = None
    if xlsx_data.get('published_coef_main') is not None:
        try:
            published_coef = float(xlsx_data['published_coef_main'])
        except (ValueError, TypeError):
            pass
    if published_coef is None and pdf_paths:
        published_coef, published_significance = _extract_pdf_coef(pdf_paths[0])
        if published_coef is not None:
            partial.setdefault('flags', []).append(
                'published_coef extracted from PDF — verify manually'
            )

    # Assemble full PaperSpec
    spec: PaperSpec = {
        # Identity
        'paper_id': paper_id,
        'paper_dir': str(paper_dir),
        'title': xlsx_data.get('paper_short_name') or paper_id,
        'source_do_file': partial.get('source_do_file'),
        'source_data_file': str(src_dir / data_file) if (src_dir and data_file) else None,
        'source_r_file': partial.get('source_r_file'),
        'replication_code_type': partial.get('replication_code_type', 'none'),
        # Regression spec
        'estimator': partial.get('estimator', ''),
        'dependent_var': partial.get('dependent_var', ''),
        'key_independent_vars': partial.get('key_independent_vars', []),
        'control_vars': partial.get('control_vars', []),
        'fixed_effects': partial.get('fixed_effects', []),
        'cluster_var': partial.get('cluster_var'),
        'sample_restrictions': partial.get('sample_restrictions', []),
        'interaction_terms': partial.get('interaction_terms', []),
        'instrumental_vars': partial.get('instrumental_vars', []),
        # Published result
        'published_coef': published_coef,
        'published_se': None,
        'published_significance': published_significance,
        # Audit
        'parse_confidence': partial.get('parse_confidence', 'low'),
        'flags': partial.get('flags', []),
        'manual_review_required': partial.get('manual_review_required', True),
        'raw_regression_command': partial.get('raw_regression_command'),
    }

    _save_spec(spec, paper_dir)
    return spec


# ---------------------------------------------------------------------------
# Backward-compat wrapper used by orchestrator.py
# ---------------------------------------------------------------------------

def parse_do_file(do_path: str, pdf_path: Optional[str] = None) -> dict:
    """Thin wrapper over _parse_do_file for backward compatibility with orchestrator.py.

    Parameters
    ----------
    do_path:
        Absolute path to the Stata ``.do`` replication file.
    pdf_path:
        Optional path to a PDF codebook (used for coefficient extraction).

    Returns
    -------
    dict
        Partial spec dict (subset of PaperSpec keys).
    """
    p = Path(do_path)
    if not p.exists():
        raise FileNotFoundError(f"DO file not found: {do_path}")
    partial = _parse_do_file(p)
    if pdf_path:
        pdf_p = Path(pdf_path)
        if pdf_p.exists():
            coef, sig = _extract_pdf_coef(pdf_p)
            if coef is not None:
                partial['published_coef'] = coef
                partial['published_significance'] = sig
    return partial


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _get_all_paper_ids(papers_root: Path) -> list[str]:
    return sorted(
        d.name for d in papers_root.iterdir()
        if d.is_dir() and (d / 'paper_info.xlsx').exists()
    )


def _print_summary_table(results: list[PaperSpec]) -> None:
    header = f"{'paper_id':<25} | {'confidence':<10} | {'estimator':<8} | {'DV':<20} | flags"
    print(header)
    print('-' * len(header))
    for spec in results:
        paper_id = spec['paper_id'][-24:]
        conf = spec['parse_confidence']
        est = spec['estimator'][:7]
        dv = spec['dependent_var'][:19]
        flags_str = spec['flags'][0][:40] if spec['flags'] else '—'
        print(f"{paper_id:<25} | {conf:<10} | {est:<8} | {dv:<20} | {flags_str}")


def main() -> None:
    parser = argparse.ArgumentParser(description='Parser Agent — extract regression spec from DO/R files')
    parser.add_argument('paper_id', nargs='?', help='Paper directory name, e.g. Paper_Meyer2024')
    parser.add_argument('--all', action='store_true', help='Parse all papers and print summary table')
    parser.add_argument('--papers-root', default=None, help='Path to papers/ directory')
    parser.add_argument('--ra-task-root', default=None, help='Path to "RA Missing data task/"')
    args = parser.parse_args()

    script_dir = Path(__file__).parent
    repo_root = script_dir.parent

    papers_root = Path(args.papers_root) if args.papers_root else repo_root / 'papers'
    ra_task_root = Path(args.ra_task_root) if args.ra_task_root else repo_root / 'RA Missing data task'

    if args.all:
        paper_ids = _get_all_paper_ids(papers_root)
        results: list[PaperSpec] = []
        for pid in paper_ids:
            print(f"  Parsing {pid}...", end=' ', flush=True)
            try:
                spec = parse_paper(pid, papers_root, ra_task_root)
                results.append(spec)
                print(f"[{spec['parse_confidence']}]")
            except Exception as e:
                print(f"[ERROR: {e}]")
        print()
        _print_summary_table(results)
    elif args.paper_id:
        spec = parse_paper(args.paper_id, papers_root, ra_task_root)
        print(json.dumps(spec, indent=2))
    else:
        parser.print_help()
        sys.exit(1)


if __name__ == '__main__':
    main()
