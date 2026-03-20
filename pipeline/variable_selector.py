"""
pipeline/variable_selector.py
------------------------------
Variable Selector: given a PaperSpec + baseline DataFrame, selects 3–5 key variables
(continuous/ordinal, non-negative preferred) and 1 aux variable to drive the power-law
MAR mechanism.  Fires Gate 2 before writing final output.

MAR mechanism: P(missing_i) ∝ aux_i ^ 1.5  (PI instructions, power-law).
CONTEXT.md sigmoid formula is incorrect — see plan Gap 1.

Usage
-----
    python pipeline/variable_selector.py Paper_Meyer2024
    python pipeline/variable_selector.py --all
    PIPELINE_ENV=test python pipeline/variable_selector.py --all --no-gate
"""

from __future__ import annotations

import argparse
import json
import logging
import math
import os
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Literal, Optional

import numpy as np
import pandas as pd

try:
    import openpyxl
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

from pipeline.config import MIN_KEY_VARS, MAX_KEY_VARS


logger = logging.getLogger(__name__)

# ── Custom exception ──────────────────────────────────────────────────────────

class PipelineDependencyError(RuntimeError):
    """Raised when a required upstream artifact is missing."""


# ── TypedDict ─────────────────────────────────────────────────────────────────

from typing import TypedDict


class VariableSelection(TypedDict):
    paper_id:             str
    key_vars:             list[str]
    aux_var:              str
    aux_var_description:  str
    key_var_rationale:    dict[str, str]
    aux_var_rationale:    str
    excluded_vars:        dict[str, str]
    selection_confidence: str   # "auto" | "partial" | "manual"
    flags:                list[str]
    human_confirmed:      bool
    correlation_matrix:   dict


# ── ID/time name set ──────────────────────────────────────────────────────────

_ID_TIME_NAMES: set[str] = {
    "id", "year", "time", "date", "fyear", "gvkey", "cusip", "permno",
    "month", "quarter", "t", "ym", "round", "qtr", "wave", "period",
    "metaid", "monthtime", "state_id", "firm_id", "company_id",
    "province", "county", "region", "district", "municipality",
    "state", "country", "census_region",
}

_ID_SUFFIXES  = ("_id", "_year", "_date", "_time", "_code", "_region", "_province")
_ID_PREFIXES  = ("id_",)


def _is_id_time(name: str) -> bool:
    low = name.lower()
    if low in _ID_TIME_NAMES:
        return True
    for suf in _ID_SUFFIXES:
        if low.endswith(suf):
            return True
    for pre in _ID_PREFIXES:
        if low.startswith(pre):
            return True
    return False


# ── Eligibility ───────────────────────────────────────────────────────────────

def _is_malformed(name: str) -> bool:
    return bool(re.search(r'[*()\[\]]', name)) or not name.isidentifier()


def _filter_eligible_vars(
    spec: dict,
    df: Optional[pd.DataFrame],
) -> tuple[list[str], dict[str, str]]:
    """Return (eligible_vars, excluded_vars_with_reasons).

    Candidate pool: key_independent_vars + control_vars.
    Data-based rules (5–10) are skipped when df is None.
    """
    excluded: dict[str, str] = {}
    eligible: list[str]      = []

    dv          = spec.get("dependent_var") or ""
    fe_set      = set(spec.get("fixed_effects")      or [])
    iv_set      = set(spec.get("instrumental_vars")  or [])
    key_iv_set  = set(spec.get("key_independent_vars") or [])

    raw_candidates: list[str] = (
        list(spec.get("key_independent_vars") or [])
        + list(spec.get("control_vars") or [])
    )

    # Expand wildcards when data available
    expanded: list[str] = []
    for c in raw_candidates:
        if not c:
            continue
        if "*" in c:
            if df is not None:
                pat = c.replace("*", ".*")
                matched = [col for col in df.columns if re.fullmatch(pat, col)]
                if matched:
                    expanded.extend(matched)
                else:
                    excluded[c] = "wildcard — no matching columns in data"
            else:
                excluded[c] = "wildcard — cannot expand without data"
        else:
            expanded.append(c)

    seen: set[str] = set()
    for v in expanded:
        if not v or v in seen:
            continue
        seen.add(v)

        # Rule 9: malformed
        if _is_malformed(v):
            excluded[v] = "malformed variable name (wildcard or non-identifier chars)"
            continue

        # Rule 1: dependent variable
        if v == dv:
            excluded[v] = "is the dependent variable"
            continue

        # Rule 2: fixed effects
        if v in fe_set:
            excluded[v] = "is a fixed-effect identifier"
            continue

        # Rule 3: instrumental variable
        if v in iv_set:
            excluded[v] = "is an instrumental variable"
            continue

        # Rule 4: ID/time name pattern
        if _is_id_time(v):
            excluded[v] = "matches ID/time name pattern (rule 4)"
            continue

        # Data-based rules — skip if no data
        if df is not None:
            # Rule 0: variable not in baseline columns (cannot validate or use)
            if v not in df.columns:
                excluded[v] = "not found in baseline data columns"
                continue

            # Rule 7: object/string dtype
            if v in df.columns and df[v].dtype == object:
                excluded[v] = "object/string dtype"
                continue

            # Rule 8: >50% missing
            if v in df.columns and df[v].isna().mean() > 0.50:
                excluded[v] = f">50% missing ({df[v].isna().mean():.0%})"
                continue

            # Rule 6: high-uniqueness identifier (integer/object dtypes only;
            # continuous floats are intentionally excluded from this check)
            if v in df.columns and df[v].dtype.kind in ("i", "u", "O"):
                n_unique = df[v].nunique()
                if n_unique / len(df) > 0.90 and n_unique > 50:
                    excluded[v] = f"high-uniqueness identifier ({n_unique} unique values)"
                    continue

            # Rule 5: binary dummy (exactly 2 unique non-NA values in {0,1})
            if v in df.columns:
                non_na = df[v].dropna()
                unique_vals = set(non_na.unique())
                if len(unique_vals) == 2 and unique_vals <= {0, 1, 0.0, 1.0}:
                    excluded[v] = "binary dummy (exactly 2 unique values in {0,1})"
                    continue

            # Rule 10: all non-positive (power law requires non-negative aux)
            if v in df.columns:
                col_vals = df[v].dropna()
                if len(col_vals) > 0 and (col_vals <= 0).all():
                    excluded[v] = "all values ≤ 0 (power-law MAR requires non-negative)"
                    continue

        eligible.append(v)

    return eligible, excluded


# ── Scoring ───────────────────────────────────────────────────────────────────

def _score_and_rank(
    eligible_vars: list[str],
    spec: dict,
    df: Optional[pd.DataFrame],
) -> list[tuple[str, float, str]]:
    """Return [(varname, score, rationale)] sorted descending by (score, -missing_rate)."""
    key_iv_set = set(spec.get("key_independent_vars") or [])
    rows: list[tuple[str, float, float, str]] = []  # (var, score, missing, rationale)

    for v in eligible_vars:
        if v in key_iv_set:
            score  = 3.0
            source = "key IV"
            missing_rate = 0.0
            if df is not None and v in df.columns:
                missing_rate = df[v].isna().mean()
            rationale = f"source: key IV; missing: {missing_rate:.1%}"
        else:
            if df is not None and v in df.columns:
                missing_rate = df[v].isna().mean()
                if missing_rate < 0.05:
                    score = 2.0
                elif missing_rate <= 0.20:
                    score = 1.0
                else:
                    score = 0.5
            else:
                score        = 2.0
                missing_rate = 0.0
            rationale = f"source: control; missing: {missing_rate:.1%}"

        non_neg = ""
        if df is not None and v in df.columns:
            all_non_neg = (df[v].dropna() >= 0).all()
            non_neg = " ≥0: Y" if all_non_neg else " ≥0: N"
        rationale += non_neg
        rows.append((v, score, missing_rate, rationale))

    rows.sort(key=lambda r: (-r[1], r[2]))
    return [(v, score, rat) for v, score, _, rat in rows]


# ── Aux variable selection ────────────────────────────────────────────────────

def _build_corr_matrix(
    vars_: list[str],
    df: Optional[pd.DataFrame],
) -> dict[str, dict[str, float]]:
    if df is None or not vars_:
        return {}
    avail = [v for v in vars_ if v in df.columns]
    if len(avail) < 2:
        return {}
    sub = df[avail].dropna()
    if len(sub) < 2:
        return {}
    corr = sub.corr(numeric_only=True)
    return {
        v: {w: round(float(corr.loc[v, w]), 4) for w in avail if w in corr.columns}
        for v in avail
        if v in corr.index
    }


def _select_aux_var(
    key_vars: list[str],
    eligible_vars: list[str],
    df: Optional[pd.DataFrame],
    threshold: float = 0.10,
) -> tuple[str, str, dict, list[str]]:
    """Return (aux_var_name, rationale, correlation_matrix, new_flags).

    Prefers non-negative vars for power-law compatibility.
    """
    new_flags: list[str] = []

    # Power-law MAR flag always added
    mar_flag = (
        "MAR mechanism: power-law P(missing) ∝ aux^1.5 (PI instructions). "
        "CONTEXT.md sigmoid formula is incorrect — see plan Gap 1."
    )
    new_flags.append(mar_flag)

    # Pool: eligible vars not in key_vars
    key_set = set(key_vars)
    pool = [v for v in eligible_vars if v not in key_set]

    if df is None:
        # No data — cannot compute correlations
        if pool:
            aux_var = pool[0]
            rationale = "selected first available (no data for correlation check)"
            corr_matrix = {}
        else:
            aux_var   = ""
            rationale = "no candidate aux vars available"
            corr_matrix = {}
        return aux_var, rationale, corr_matrix, new_flags

    # Filter: missing_rate < 5%
    pool = [v for v in pool if v in df.columns and df[v].isna().mean() < 0.05]

    # Prefer non-negative vars
    non_neg_pool = [
        v for v in pool
        if (df[v].dropna() >= 0).all()
    ]
    if non_neg_pool:
        search_pool = non_neg_pool
    else:
        search_pool = pool
        if pool:
            new_flags.append(
                "aux var may be negative — power-law MAR formula requires non-negative "
                "values; consider log transform"
            )

    # Compute correlations with key_vars
    def _majority_correlated(v: str, thresh: float) -> tuple[bool, float]:
        if not key_vars:
            return False, 0.0
        rs = []
        for kv in key_vars:
            if kv not in df.columns or v not in df.columns:
                continue
            sub = df[[kv, v]].dropna()
            if len(sub) < 3:
                continue
            r = sub[kv].corr(sub[v])
            if not np.isnan(r):
                rs.append(abs(r))
        if not rs:
            return False, 0.0
        n_corr = sum(1 for r in rs if r >= thresh)
        majority = n_corr >= math.ceil(len(key_vars) / 2)
        return majority, float(np.mean(rs))

    best_aux   = ""
    best_mean_r = -1.0

    for v in search_pool:
        majority, mean_r = _majority_correlated(v, threshold)
        if majority and mean_r > best_mean_r:
            best_aux    = v
            best_mean_r = mean_r

    # Relax threshold if nothing found
    if not best_aux and threshold > 0.05:
        new_flags.append("aux var correlation relaxed to 0.05")
        for v in search_pool:
            majority, mean_r = _majority_correlated(v, 0.05)
            if majority and mean_r > best_mean_r:
                best_aux    = v
                best_mean_r = mean_r

    if not best_aux:
        new_flags.append("no suitable aux var found — human must select")
        return "", "no suitable aux var found", {}, new_flags

    # Build correlation matrix for key_vars + aux_var
    all_vars = [v for v in (key_vars + [best_aux]) if v]
    corr_matrix = _build_corr_matrix(all_vars, df)

    rationale = (
        f"mean |r| with key vars: {best_mean_r:.3f}; "
        f"non-negative: {(df[best_aux].dropna() >= 0).all()}"
    )
    return best_aux, rationale, corr_matrix, new_flags


def _pick_aux_first(
    aux_pool: list[str],
    all_eligible: list[str],
    df: Optional[pd.DataFrame],
    threshold: float = 0.10,
) -> tuple[str, str, dict, list[str]]:
    """Reserve aux_var from aux_pool before key_var selection.

    aux_pool    — eligible vars allowed to become aux (key IVs excluded by caller)
    all_eligible — full eligible pool (used for correlation context + matrix)
    """
    new_flags: list[str] = []
    new_flags.append(
        "MAR mechanism: power-law P(missing) ∝ aux^1.5 (PI instructions). "
        "CONTEXT.md sigmoid formula is incorrect — see plan Gap 1."
    )

    if df is None:
        if aux_pool:
            return aux_pool[0], "selected first available (no data for correlation check)", {}, new_flags
        return "", "no eligible aux candidates", {}, new_flags

    # Filter: < 20% missing (relaxed from 5% to avoid empty pool on real datasets)
    candidates = [v for v in aux_pool if v in df.columns and df[v].isna().mean() < 0.20]

    # Prefer non-negative (power-law compatibility)
    non_neg = [v for v in candidates if (df[v].dropna() >= 0).all()]
    search_pool = non_neg if non_neg else candidates
    if candidates and not non_neg:
        new_flags.append(
            "aux var may be negative — power-law MAR formula requires non-negative "
            "values; consider log transform"
        )

    def _mean_r_with_others(v: str, thresh: float) -> tuple[bool, float]:
        others = [w for w in all_eligible if w != v and w in df.columns]
        if not others:           # no other vars to correlate with → accept
            return True, 0.0
        rs = []
        for w in others:
            sub = df[[v, w]].dropna()
            if len(sub) < 3:
                continue
            r = sub[v].corr(sub[w])
            if not np.isnan(r):
                rs.append(abs(r))
        if not rs:
            return False, 0.0
        n_corr  = sum(1 for r in rs if r >= thresh)
        majority = n_corr >= math.ceil(len(others) / 2)
        return majority, float(np.mean(rs))

    best_aux, best_mean_r = "", -1.0
    for v in search_pool:
        majority, mean_r = _mean_r_with_others(v, threshold)
        if majority and mean_r > best_mean_r:
            best_aux, best_mean_r = v, mean_r

    # Threshold relaxation
    if not best_aux and threshold > 0.05:
        new_flags.append("aux var correlation relaxed to 0.05")
        for v in search_pool:
            majority, mean_r = _mean_r_with_others(v, 0.05)
            if majority and mean_r > best_mean_r:
                best_aux, best_mean_r = v, mean_r

    if not best_aux and search_pool:
        best_aux = min(
            search_pool,
            key=lambda v: df[v].isna().mean() if v in df.columns else 1.0,
        )
        new_flags.append(
            f"aux var {best_aux!r} selected by lowest missing rate "
            "(no correlation threshold met — human review recommended)"
        )

    if not best_aux:
        new_flags.append("no suitable aux var found — human must select")
        return "", "no suitable aux var found", {}, new_flags

    corr_matrix = _build_corr_matrix([v for v in all_eligible if v], df)
    rationale = (
        f"mean |r| with other eligible vars: {best_mean_r:.3f}; "
        f"non-negative: {(df[best_aux].dropna() >= 0).all()}"
    )
    return best_aux, rationale, corr_matrix, new_flags


# ── Confidence ────────────────────────────────────────────────────────────────

def _determine_confidence(
    spec: dict,
    eligible_vars: list[str],
    aux_var: str,
    data_available: bool,
) -> str:
    if (
        spec.get("replication_code_type") == "none"
        or spec.get("parse_confidence") == "low"
        or len(eligible_vars) < MIN_KEY_VARS
    ):
        return "manual"
    if (
        spec.get("parse_confidence") == "medium"
        or not data_available
        or not aux_var
    ):
        return "partial"
    return "auto"


# ── Log helper ────────────────────────────────────────────────────────────────

def _append_log(paper_id: str, message: str) -> None:
    log_dir = Path(__file__).parent.parent / "logs"
    log_dir.mkdir(exist_ok=True)
    log_path = log_dir / f"{paper_id}_log.md"
    with open(log_path, "a", encoding="utf-8") as f:
        ts = datetime.now().isoformat(timespec="seconds")
        f.write(f"\n## [{ts}] variable_selector\n")
        f.write(message + "\n")


# ── Save selection ────────────────────────────────────────────────────────────

def _save_selection(selection: dict, papers_root: str) -> None:
    """Write selection.json, update paper_info.xlsx, append to log."""
    paper_id  = selection["paper_id"]
    paper_dir = Path(papers_root) / paper_id
    paper_dir.mkdir(parents=True, exist_ok=True)

    # 1. Write selection.json
    sel_path = paper_dir / "selection.json"
    # Make a JSON-serialisable copy (remove any non-serialisable objects)
    sel_copy = {k: v for k, v in selection.items()}
    with open(sel_path, "w", encoding="utf-8") as f:
        json.dump(sel_copy, f, indent=2)

    # 2. Update paper_info.xlsx (if it exists)
    xlsx_path = paper_dir / "paper_info.xlsx"
    if xlsx_path.exists() and _HAS_OPENPYXL:
        wb = openpyxl.load_workbook(str(xlsx_path))
        ws = wb.active
        # Find data row matching paper_id
        data_row: Optional[int] = None
        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[0].value == paper_id:
                data_row = row[0].row
                break
        if data_row is None:
            data_row = 2  # fallback

        # col 12: key_vars, col 13: aux_var — {name} — {description}
        ws.cell(row=data_row, column=12).value = ", ".join(selection.get("key_vars") or [])
        aux_name = selection.get("aux_var") or ""
        aux_desc = selection.get("aux_var_description") or ""
        aux_cell = f"{aux_name} — {aux_desc}" if aux_desc else aux_name
        ws.cell(row=data_row, column=13).value = aux_cell
        wb.save(str(xlsx_path))

    # 3. Append to log
    _append_log(
        paper_id,
        f"selection saved: key_vars={selection.get('key_vars')}, "
        f"aux_var={selection.get('aux_var')!r}",
    )


# ── Human gate display ────────────────────────────────────────────────────────

_WIDTH = 64
_LINE  = "═" * _WIDTH


def _box_line(text: str = "") -> str:
    return f"║  {text:<{_WIDTH - 4}}║"


def _print_box(lines: list[str]) -> None:
    print(f"╔{_LINE}╗")
    for line in lines:
        if line == "---":
            print(f"╠{_LINE}╣")
        else:
            print(_box_line(line))
    print(f"╚{_LINE}╝")


def _human_gate(
    selection: dict,
    paper_id: str,
    auto_confirm: bool = False,
) -> dict:
    """Display summary → prompt [C]onfirm / [E]dit / [A]bort."""
    df_ref: Optional[pd.DataFrame] = selection.get("_df_ref")  # internal ref for re-validation
    spec_ref: Optional[dict]       = selection.get("_spec_ref")

    while True:
        key_vars    = selection.get("key_vars") or []
        aux_var     = selection.get("aux_var") or ""
        confidence  = selection.get("selection_confidence", "?")
        flags       = selection.get("flags") or []
        excl        = selection.get("excluded_vars") or {}
        kv_rat      = selection.get("key_var_rationale") or {}
        aux_rat     = selection.get("aux_var_rationale") or ""
        corr_matrix = selection.get("correlation_matrix") or {}

        lines: list[str] = [
            f"VARIABLE SELECTOR — Human Gate: {paper_id}",
            "---",
            f"Selection confidence: {confidence}",
            "MAR mechanism: power-law  P(missing) ∝ aux^1.5",
            "---",
            f"KEY VARIABLES ({len(key_vars)} of {MIN_KEY_VARS}–{MAX_KEY_VARS}):",
        ]
        for i, v in enumerate(key_vars, 1):
            rat = kv_rat.get(v, "")
            lines.append(f"  {i}. {v:<22}  {rat}")
        lines.append("---")
        lines.append("AUX VARIABLE PROPOSED:")
        if aux_var:
            lines.append(f"  {aux_var}")
            if aux_rat:
                lines.append(f"  {aux_rat}")
            # Per-key-var correlations
            aux_corrs = corr_matrix.get(aux_var, {})
            for kv in key_vars:
                r = aux_corrs.get(kv, aux_corrs.get(kv))
                if r is not None:
                    lines.append(f"    {kv}: r = {r:.4f}")
        else:
            lines.append("  (none — human must select)")
        if excl:
            lines.append("---")
            lines.append(f"EXCLUDED ({len(excl)} vars):")
            for v, reason in list(excl.items())[:10]:
                lines.append(f"  {v}: {reason}")
            if len(excl) > 10:
                lines.append(f"  ... and {len(excl) - 10} more")
        if flags:
            lines.append("---")
            lines.append("FLAGS:")
            for fl in flags:
                lines.append(f"  ⚠  {fl}")
        lines.append("---")
        lines.append("[C]onfirm  [E]dit  [A]bort")

        if not auto_confirm:
            _print_box(lines)
        else:
            logger.debug("[%s] variable_selector auto-confirming: key_vars=%s aux_var=%r",
                         paper_id,
                         selection.get("key_vars"),
                         selection.get("aux_var"))

        if auto_confirm:
            response = "c"
        else:
            response = input("Choice: ").strip().lower()

        if response in ("c", "confirm", ""):
            result = {**selection, "human_confirmed": True}
            result.pop("_df_ref", None)
            result.pop("_spec_ref", None)
            return result

        elif response in ("e", "edit"):
            print("  Current key_vars:", key_vars)
            new_kv_raw = input("  New key_vars (comma-separated, blank=keep): ").strip()
            if new_kv_raw:
                new_kv = [v.strip() for v in new_kv_raw.split(",") if v.strip()]
                selection = {**selection, "key_vars": new_kv}

            print(f"  Current aux_var: {aux_var!r}")
            new_aux = input("  New aux_var name (blank=keep): ").strip()
            if new_aux:
                selection = {**selection, "aux_var": new_aux}

            new_desc = input("  One-line description of aux_var: ").strip()
            if new_desc:
                selection = {**selection, "aux_var_description": new_desc}

            # Re-validate key_vars if spec/df available
            if spec_ref is not None:
                _elig, _excl = _filter_eligible_vars(spec_ref, df_ref)
                invalid = [v for v in (selection.get("key_vars") or []) if v not in _elig]
                if invalid:
                    print(f"  ⚠  Not in eligible pool: {invalid}. Keeping anyway.")

        elif response in ("a", "abort"):
            raise SystemExit(f"Variable selection aborted for {paper_id}.")

        else:
            print("  Please enter C, E, or A.")


# ── Public entry ──────────────────────────────────────────────────────────────

def select_variables(
    paper_id: str,
    papers_root: str,
    spec: Optional[dict] = None,
    data: Optional[pd.DataFrame] = None,
    auto_confirm: bool = False,
) -> dict:
    """Select key_vars and aux_var for the MAR simulation.

    Parameters
    ----------
    paper_id:
        Paper identifier (e.g. 'Paper_Meyer2024').
    papers_root:
        Absolute path to the papers/ root directory.
    spec:
        PaperSpec dict; loaded from papers/{paper_id}/spec.json if None.
    data:
        Baseline DataFrame; loaded from papers/{paper_id}/baseline.parquet if None.
    auto_confirm:
        Skip interactive gate (for testing / PIPELINE_ENV=test).

    Returns
    -------
    VariableSelection dict.
    """
    # ── Gap 2: dependency guard ────────────────────────────────────────────────
    if data is None:
        baseline_path = Path(papers_root) / paper_id / "baseline.parquet"
        if not baseline_path.exists():
            raise PipelineDependencyError(
                f"[{paper_id}] baseline.parquet not found at {baseline_path}. "
                "Run data_prep_agent + baseline_verifier (and approve Gate 1) first."
            )
        data = pd.read_parquet(str(baseline_path))

    # ── Load spec ──────────────────────────────────────────────────────────────
    if spec is None:
        spec_path = Path(papers_root) / paper_id / "spec.json"
        if not spec_path.exists():
            raise PipelineDependencyError(
                f"[{paper_id}] spec.json not found at {spec_path}. "
                "Run parser_agent first."
            )
        with open(spec_path, encoding="utf-8") as f:
            spec = json.load(f)

    data_available = data is not None

    # ── Eligibility ────────────────────────────────────────────────────────────
    eligible_vars, excluded_vars = _filter_eligible_vars(spec, data)
    flags: list[str] = []

    # ── Step A: reserve aux_var first (key IVs protected from aux pool) ────────
    key_iv_set = set(spec.get("key_independent_vars") or [])
    aux_pool = [v for v in eligible_vars if v not in key_iv_set] or eligible_vars
    aux_var, aux_rationale, corr_matrix, aux_flags = _pick_aux_first(
        aux_pool, eligible_vars, data
    )
    flags.extend(aux_flags)

    # ── Step B: select key_vars from remaining eligible pool ───────────────────
    key_pool = [v for v in eligible_vars if v != aux_var]
    ranked = _score_and_rank(key_pool, spec, data)

    if len(ranked) < MIN_KEY_VARS:
        _aux_label = repr(aux_var) if aux_var else "none"
        flags.append(
            f"only {len(ranked)} distinct variable(s) remain for key_vars "
            f"(after reserving aux_var={_aux_label}) — "
            f"need at least {MIN_KEY_VARS}; human must select or paper may be infeasible"
        )
        key_vars: list[str] = []
        key_var_rationale: dict[str, str] = {}
    else:
        selected_ranked = ranked[:MAX_KEY_VARS]
        key_vars = [v for v, _, _ in selected_ranked]
        key_var_rationale = {v: rat for v, _, rat in selected_ranked}

    # ── Post-selection validation of key_vars against baseline data ────────────
    if data is not None and key_vars:
        invalid_kv = [
            v for v in key_vars
            if v not in data.columns
            or data[v].dtype == object
            or set(data[v].dropna().unique()) <= {0, 1, 0.0, 1.0}
        ]
        if invalid_kv:
            flags.append(f"key_vars failed post-selection validation: {invalid_kv}")
            key_vars = [v for v in key_vars if v not in invalid_kv]
            key_var_rationale = {v: r for v, r in key_var_rationale.items() if v in key_vars}

    if len(key_vars) < MIN_KEY_VARS:
        flags.append(
            f"INFEASIBLE: only {len(key_vars)} valid key_var(s) remain "
            f"after baseline validation (need ≥{MIN_KEY_VARS}). "
            "Paper requires manual specification — cannot run MAR simulation."
        )

    # ── Confidence ─────────────────────────────────────────────────────────────
    confidence = _determine_confidence(spec, ranked, aux_var, data_available)

    # ── Draft selection ────────────────────────────────────────────────────────
    draft: dict = {
        "paper_id":             paper_id,
        "key_vars":             key_vars,
        "aux_var":              aux_var,
        "aux_var_description":  "",
        "key_var_rationale":    key_var_rationale,
        "aux_var_rationale":    aux_rationale,
        "excluded_vars":        excluded_vars,
        "selection_confidence": confidence,
        "flags":                flags,
        "human_confirmed":      False,
        "correlation_matrix":   corr_matrix,
        # internal refs for edit-loop re-validation (stripped before return)
        "_df_ref":   data,
        "_spec_ref": spec,
    }

    # ── Human gate (Gate 2) ────────────────────────────────────────────────────
    confirmed = _human_gate(draft, paper_id, auto_confirm=auto_confirm)

    # ── Strip internal refs (not serialisable, not part of VariableSelection) ──
    clean = {k: v for k, v in confirmed.items() if not k.startswith("_")}

    # ── Persist ────────────────────────────────────────────────────────────────
    if clean.get("human_confirmed"):
        _save_selection(clean, papers_root)

    return clean


# ── CLI ───────────────────────────────────────────────────────────────────────

def _find_papers_root() -> Path:
    here = Path(__file__).parent.parent
    papers = here / "papers"
    if papers.is_dir():
        return papers
    raise FileNotFoundError(f"papers/ directory not found under {here}")


def _run_paper_cli(paper_id: str, papers_root: Path, no_gate: bool) -> None:
    try:
        selection = select_variables(
            paper_id,
            str(papers_root),
            auto_confirm=no_gate,
        )
        print(f"[{paper_id}] key_vars: {selection.get('key_vars')}")
        print(f"[{paper_id}] aux_var:  {selection.get('aux_var')!r}")
        print(f"[{paper_id}] confidence: {selection.get('selection_confidence')}")
    except PipelineDependencyError as exc:
        print(f"[{paper_id}] DEPENDENCY ERROR: {exc}", file=sys.stderr)
    except SystemExit as exc:
        print(f"[{paper_id}] ABORTED: {exc}", file=sys.stderr)
    except Exception as exc:
        print(f"[{paper_id}] ERROR: {exc}", file=sys.stderr)


def main() -> None:
    parser = argparse.ArgumentParser(description="Variable Selector")
    parser.add_argument("paper_id", nargs="?", help="Paper ID")
    parser.add_argument("--all",     action="store_true", help="Run for all papers")
    parser.add_argument("--no-gate", action="store_true",
                        help="Skip human gate (PIPELINE_ENV=test only)")
    args = parser.parse_args()

    if args.no_gate and os.environ.get("PIPELINE_ENV") != "test":
        sys.exit("--no-gate is only allowed when PIPELINE_ENV=test")

    papers_root = _find_papers_root()

    if args.all:
        for pd_dir in sorted(papers_root.glob("Paper_*/")):
            _run_paper_cli(pd_dir.name, papers_root, no_gate=args.no_gate)
    elif args.paper_id:
        _run_paper_cli(args.paper_id, papers_root, no_gate=args.no_gate)
    else:
        parser.print_help()
        sys.exit(1)


if __name__ == "__main__":
    main()
